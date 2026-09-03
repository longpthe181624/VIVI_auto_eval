import os
import sys
import re
import json
import gc
import asyncio
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, List, Tuple

# Ensure project root and site-packages are in sys.path
project_root = Path(__file__).resolve().parent.parent
site_packages = project_root / "myenv" / "lib" / "python3.14" / "site-packages"
if site_packages.exists() and str(site_packages) not in sys.path:
    sys.path.insert(0, str(site_packages))
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import openpyxl
from openpyxl.styles import Alignment
import src.config as config
from src.eval_tools import rag_rule_search, eval_test_result, web_search_verification
from src.test_eval_agent import TestEvalAgent
from src.report_builder import (
    EXTENDED_HEADERS, build_extended_row, style_header_row, style_data_row,
    autosize_columns, build_summary_sheet, extract_domain, SEVERITY_FILLS,
)


def _detect_columns(header_row: List[Any]) -> Dict[str, int]:
    """Dynamically maps Excel column headers to standard evaluation fields with strict priority."""
    col_map = {}
    if not header_row:
        return col_map

    header_strs = [str(c).strip().lower() if c is not None else "" for c in header_row]

    for idx, h in enumerate(header_strs):
        if not h:
            continue

        if any(k in h for k in ["auto_eval_result", "auto_eval"]):
            col_map.setdefault("auto_eval_start", idx)

        # 1. Test Case Name
        elif any(k in h for k in ["name_testcase", "testcase_name", "tc_name", "name testcase", "testcase", "test_case", "tc"]):
            col_map.setdefault("name", idx)
        elif h in ["name", "id", "stt"]:
            col_map.setdefault("name", idx)

        # 2. User Command / Prompt
        elif any(k in h for k in ["user_command", "user command", "câu_lệnh", "câu lệnh", "prompt", "query", "question"]):
            col_map.setdefault("user_command", idx)
        elif h in ["command", "lệnh"]:
            col_map.setdefault("user_command", idx)

        # 3. Vivi Listen (STT Transcribed Input)
        elif any(k in h for k in ["vivi_listen", "vivi listen", "listen", "transcribed", "nghe"]):
            col_map.setdefault("vivi_listen", idx)

        # 4. Actual Response
        elif any(k in h for k in ["actual_resp", "actual_response", "actual response", "actual", "model_answer", "phản_hồi_thực_tế", "phản hồi thực tế", "vivi_response", "vivi response"]):
            col_map.setdefault("actual_resp", idx)
        elif h in ["phản hồi", "phản_hồi"]:
            col_map.setdefault("actual_resp", idx)

        # 5. Expected Response / Ground Truth
        elif any(k in h for k in ["expected_resp", "expected_response", "expected response", "expected", "ground_truth", "kết_quả_mong_muốn", "kết quả mong muốn", "target"]):
            col_map.setdefault("expected_resp", idx)
        elif h in ["kỳ vọng", "kỳ_vọng"]:
            col_map.setdefault("expected_resp", idx)

        # 6. Previous Result Status
        elif h in ["result", "status", "trạng_thái", "kết quả", "kết_quả"]:
            col_map.setdefault("prev_result", idx)

        # 7. Latency / Time
        elif any(k in h for k in ["latency", "time", "duration", "thời gian"]):
            col_map.setdefault("latency", idx)

    return col_map


def trim_trailing_empty(row_values: tuple) -> list:
    """Trims trailing None or empty whitespace values from an Excel row tuple dynamically."""
    if not row_values:
        return []
    last_idx = -1
    for idx in range(len(row_values) - 1, -1, -1):
        v = row_values[idx]
        if v is not None and str(v).strip() != "":
            last_idx = idx
            break
    if last_idx == -1:
        return []
    return list(row_values[:last_idx + 1])


def detect_testcase_category(file_name: str, sheet_name: str) -> str:
    """Auto-detects whether a test case workbook or sheet belongs to owner_manual, command_rule, general_knowledge, error_code, etc."""
    combined = f"{file_name} {sheet_name}".lower()
    if any(k in combined for k in ["thuongthuc", "thưởng thức", "general", "trivia", "web_search", "sensitive"]):
        return "general_knowledge"
    elif any(k in combined for k in ["command", "_va_", "va_command", "lệnh", "tính năng", "feature"]):
        return "command_rule"
    elif re.search(r"\bom\b|om\d+|manual|hướng dẫn|huong dan|chủ sở hữu", combined):
        return "owner_manual"
    elif any(k in combined for k in ["dtc", "error", "mã lỗi"]):
        return "error_code"
    return "general"


from src.eval_router import AdaptiveEvalRouter


class ExcelTestEvaluator:
    """Automated RAG-enabled Excel test case evaluator."""

    def __init__(self):
        self.eval_agent = TestEvalAgent()
        self.router = AdaptiveEvalRouter()

    def evaluate_row_sync(
        self,
        name: str,
        user_cmd: str,
        actual_resp: str,
        expected_resp: str,
        category: str = None,
        vivi_listen: str = ""
    ) -> Dict[str, Any]:
        """Evaluates a single test case row via multi-tiered adaptive routing."""
        return self.router.evaluate(name, user_cmd, actual_resp, expected_resp, testcase_category=category, vivi_listen=vivi_listen)

    def evaluate_file(self, input_excel_path: str, output_excel_path: str = None) -> str:
        """Evaluates all test cases in an Excel file using lightweight streaming mode to prevent memory bloat."""
        input_path = Path(input_excel_path)
        if not input_path.exists():
            raise FileNotFoundError(f"Input Excel file '{input_excel_path}' not found.")

        print(f"\n🧪 Starting Automated Evaluation on: {input_path.name}")
        try:
            wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to parse Excel file '{input_excel_path}': {e}")

        output_path = Path(output_excel_path) if output_excel_path else input_path.parent / f"evaluated_{input_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)  # Remove default blank sheet

        total_eval_count = 0
        pass_count = 0
        fail_count = 0
        retest_count = 0
        severity_counts: Dict[str, int] = {"PASS": 0, "LOW": 0, "MEDIUM": 0, "HIGH": 0}
        domain_stats: Dict[str, Dict[str, int]] = {}

        new_headers = EXTENDED_HEADERS

        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]
            ws_out = wb_out.create_sheet(title=sheet_name)

            header_vals = None
            header_row_num = -1
            col_map = {}
            raw_rows_buffer = []

            # Stream rows to find header and limit column bounds dynamically
            for r_idx, row in enumerate(ws_in.iter_rows(values_only=True), 1):
                trimmed_row = trim_trailing_empty(row)
                if not trimmed_row:
                    continue

                if header_vals is None:
                    test_map = _detect_columns(trimmed_row)
                    if "user_command" in test_map or "actual_resp" in test_map:
                        header_vals = [str(c).strip() if c is not None else "" for c in trimmed_row]
                        col_map = test_map
                        header_row_num = r_idx
                        continue
                    else:
                        # Copy pre-header rows (e.g. metadata/title block) directly
                        ws_out.append(trimmed_row)
                else:
                    raw_rows_buffer.append(trimmed_row)

            if not header_vals or ("user_command" not in col_map and "actual_resp" not in col_map):
                # Sheet has no test case headers
                continue

            print(f"  📊 Processing Sheet: '{sheet_name}' ({len(raw_rows_buffer)} test cases)...")

            # Determine actual valid column count
            max_data_col = max(col_map.values()) if col_map else len(header_vals) - 1
            max_col_idx = min(len(header_vals), max_data_col + 5)
            clean_header = header_vals[:max_col_idx]

            # Output header setup
            if "auto_eval_start" in col_map:
                col_start_idx = col_map["auto_eval_start"] + 1
            else:
                col_start_idx = len(clean_header) + 1

            full_header = clean_header + new_headers
            ws_out.append(full_header)
            style_header_row(ws_out, ws_out.max_row, col_start_idx, len(full_header))

            file_category = detect_testcase_category(input_path.name, sheet_name)

            from concurrent.futures import ThreadPoolExecutor

            # Process test case data rows in parallel using 16 worker threads
            def process_row_item(item):
                r_idx, row_vals = item
                name_val = str(row_vals[col_map["name"]]).strip() if "name" in col_map and col_map["name"] < len(row_vals) and row_vals[col_map["name"]] is not None else f"TC_{r_idx}"
                user_cmd = str(row_vals[col_map["user_command"]]).strip() if "user_command" in col_map and col_map["user_command"] < len(row_vals) and row_vals[col_map["user_command"]] is not None else ""
                vivi_listen = str(row_vals[col_map["vivi_listen"]]).strip() if "vivi_listen" in col_map and col_map["vivi_listen"] < len(row_vals) and row_vals[col_map["vivi_listen"]] is not None else ""
                actual_resp = str(row_vals[col_map["actual_resp"]]).strip() if "actual_resp" in col_map and col_map["actual_resp"] < len(row_vals) and row_vals[col_map["actual_resp"]] is not None else ""
                expected_resp = str(row_vals[col_map["expected_resp"]]).strip() if "expected_resp" in col_map and col_map["expected_resp"] < len(row_vals) and row_vals[col_map["expected_resp"]] is not None else ""

                res = self.evaluate_row_sync(
                    name=name_val,
                    user_cmd=user_cmd,
                    actual_resp=actual_resp,
                    expected_resp=expected_resp,
                    category=file_category,
                    vivi_listen=vivi_listen
                )
                return r_idx, row_vals, res

            items = list(enumerate(raw_rows_buffer, 1))
            with ThreadPoolExecutor(max_workers=16) as pool:
                eval_results = list(pool.map(process_row_item, items))

            for r_idx, row_vals, res in eval_results:
                status = res["auto_result"]
                severity = res.get("severity", "MEDIUM")
                total_eval_count += 1
                severity_counts[severity] = severity_counts.get(severity, 0) + 1
                if status == "PASS":
                    pass_count += 1
                elif status == "FAIL":
                    fail_count += 1
                else:
                    retest_count += 1

                name_val = str(row_vals[col_map["name"]]).strip() if "name" in col_map and col_map["name"] < len(row_vals) and row_vals[col_map["name"]] is not None else f"TC_{r_idx}"
                domain = extract_domain(name_val)
                d_stats = domain_stats.setdefault(domain, {"total": 0, "pass": 0, "fail": 0, "retest": 0})
                d_stats["total"] += 1
                d_stats["pass" if status == "PASS" else "fail" if status == "FAIL" else "retest"] += 1

                # Build row values
                base_row = list(row_vals[:max_col_idx])
                while len(base_row) < max_col_idx:
                    base_row.append("")

                out_row_vals = base_row + build_extended_row(res)
                ws_out.append(out_row_vals)

                curr_row = ws_out.max_row
                style_data_row(ws_out, curr_row, col_start_idx, len(out_row_vals), severity)

                if "prev_result" in col_map and col_map["prev_result"] < len(base_row):
                    orig_res_cell = ws_out.cell(row=curr_row, column=col_map["prev_result"] + 1, value=res["auto_result"])
                    orig_res_cell.fill = SEVERITY_FILLS.get(severity)
                    orig_res_cell.alignment = Alignment(horizontal="center")

            autosize_columns(ws_out, full_header)

        build_summary_sheet(wb_out, total_eval_count, pass_count, fail_count, retest_count, severity_counts, domain_stats)

        # Determine output filename
        if not output_excel_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_excel_path = str(input_path.parent / f"evaluated_{input_path.stem}_{timestamp}.xlsx")

        wb_out.save(output_excel_path)

        print("\n" + "=" * 60)
        print(" 🎉 AUTOMATED EVALUATION COMPLETE!")
        print("=" * 60)
        print(f" 📄 Total Test Cases Evaluated : {total_eval_count}")
        print(f" ✅ PASS                       : {pass_count}")
        print(f" ❌ FAIL                       : {fail_count}")
        print(f" ⚠️  RETEST                     : {retest_count}")
        if total_eval_count > 0:
            print(f" 📈 Pass Rate                  : {pass_count / total_eval_count * 100:.1f}%")
        print(f" 💾 Output Saved To            : {output_excel_path}")
        print("=" * 60 + "\n")

        return output_excel_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Automated Excel Test Case Evaluator")
    parser.add_argument("--file", type=str, default="data/kết quảOM8NP.xlsx", help="Input Excel test case file")
    args = parser.parse_args()

    evaluator = ExcelTestEvaluator()
    evaluator.evaluate_file(args.file)
