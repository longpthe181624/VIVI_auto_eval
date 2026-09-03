"""
Shared Excel Report Building Module.

Centralizes the extended diagnostic columns, cell styling, and summary-sheet
generation used by both the CLI batch path (excel_evaluator.py) and the web
dashboard batch path (web_server.py), so the two output formats stay in sync
instead of drifting apart (the web path previously had no cell coloring at all,
unlike the CLI path).

Adds three diagnostic columns beyond the original 5 (`Auto_Eval_Result`,
`Similarity_Score(%)`, `Matched_Rule_Spec`, `Root_Cause_Analysis`,
`Suggested_Remediation`): `Severity` (PASS/LOW/MEDIUM/HIGH), `Semantic_Error_Pct(%)`
(how wrong, quantified), and `Missing_Key_Terms` (where it's wrong - the specific
terms from the reference the answer failed to cover) plus `Error_Category` (why -
FACT_HALLUCINATION / FALSE_REFUSAL / STT_ACOUSTIC_MISMATCH / COMPLETENESS_LOSS /
INFRASTRUCTURE_OUTAGE / NONE) and `Trace_ID` for cross-referencing the full audit
trace_log via the API if deeper investigation is needed.
"""

import re
import datetime
from typing import Dict, Any, List
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

SEVERITY_FILLS = {
    "PASS": PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"),
    "LOW": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "MEDIUM": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
    "HIGH": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
}
SEVERITY_FONT_COLORS = {
    "PASS": "155724",
    "LOW": "1B4F8C",
    "MEDIUM": "856404",
    "HIGH": "721C24",
}
THIN_SIDE = Side(style="thin", color="D9D9D9")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

EXTENDED_HEADERS = [
    "Auto_Eval_Result",
    "Severity",
    "Similarity_Score(%)",
    "Semantic_Error_Pct(%)",
    "Error_Category",
    "Missing_Key_Terms",
    "Matched_Rule_Spec",
    "Root_Cause_Analysis",
    "Suggested_Remediation",
    "Trace_ID",
]

_DOMAIN_PATTERN = re.compile(r"\[[^\-\]]+-([^\-\]]+)-\d+\]")


def extract_domain(name: str) -> str:
    """Extracts the domain/category segment from a test case name like
    "[VF8-Gioithieu-0001] Verify ..." -> "Gioithieu". Mirrors the same regex
    used client-side in static/js/app.js for the dashboard's domain breakdown,
    so both stay consistent."""
    if not name:
        return "Unknown"
    m = _DOMAIN_PATTERN.match(name.strip())
    return m.group(1).strip() if m else "Unknown"


def build_extended_row(res: Dict[str, Any]) -> List[Any]:
    """Builds the extended diagnostic column values for one evaluated row,
    matching the order of EXTENDED_HEADERS."""
    trace = res.get("trace_log") or {}
    semantic_diff = trace.get("semantic_diff") or {}
    missing_terms = semantic_diff.get("missing_keywords") or []
    if missing_terms:
        missing_str = ", ".join(missing_terms[:8])
    else:
        missing_str = "" if res["auto_result"] == "PASS" else "-"

    return [
        res["auto_result"],
        res.get("severity", ""),
        res["score"],
        res.get("semantic_error_pct", ""),
        res.get("error_category", ""),
        missing_str,
        res["rule_info"],
        res["rca"],
        res["remediation"],
        res.get("trace_id", ""),
    ]


def style_header_row(ws, header_row_num: int, col_start_idx: int, total_cols: int):
    for c_idx in range(col_start_idx, total_cols + 1):
        cell = ws.cell(row=header_row_num, column=c_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_row(ws, row_num: int, col_start_idx: int, total_cols: int, severity: str):
    """Colors the verdict + severity cells by severity (HIGH=red, MEDIUM=amber,
    LOW=blue, PASS=green) and applies a light border/wrap to every extended
    column in the row for readability."""
    fill = SEVERITY_FILLS.get(severity, SEVERITY_FILLS["MEDIUM"])
    font_color = SEVERITY_FONT_COLORS.get(severity, "000000")

    for c_idx in range(col_start_idx, total_cols + 1):
        cell = ws.cell(row=row_num, column=c_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    verdict_cell = ws.cell(row=row_num, column=col_start_idx)
    verdict_cell.fill = fill
    verdict_cell.font = Font(bold=True, color=font_color)
    verdict_cell.alignment = Alignment(horizontal="center", vertical="center")

    severity_cell = ws.cell(row=row_num, column=col_start_idx + 1)
    severity_cell.fill = fill
    severity_cell.font = Font(bold=True, color=font_color)
    severity_cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, header_row: List[str], max_width: int = 60, min_width: int = 10, sample_rows: int = 60):
    """Cheap column-width heuristic: sample the first `sample_rows` data rows
    only (not the whole sheet) to keep this fast on large files."""
    for idx in range(1, len(header_row) + 1):
        col_letter = get_column_letter(idx)
        best_len = min_width
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, sample_rows), min_col=idx, max_col=idx):
            for cell in row:
                if cell.value:
                    best_len = max(best_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = best_len + 2


def build_summary_sheet(wb, total: int, pass_cnt: int, fail_cnt: int, retest_cnt: int,
                         severity_counts: Dict[str, int], domain_stats: Dict[str, Dict[str, int]]):
    """Creates a 'Summary' sheet as the first sheet: KPI cards, severity
    breakdown, and a domain/category breakdown table sorted by error rate -
    the same analysis the dashboard's domain-breakdown panel shows, so the
    exported file carries the same at-a-glance diagnosis."""
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False

    ws["B2"] = "ViVi Auto-Eval — Evaluation Summary"
    ws["B2"].font = Font(size=16, bold=True, color="1F4E78")
    ws["B3"] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["B3"].font = Font(size=9, italic=True, color="808080")

    kpis = [
        ("Total Cases", total, "1F4E78"),
        ("Pass Rate", f"{(pass_cnt / total * 100):.1f}%" if total else "0%", "1E8449"),
        ("Fail Count", fail_cnt, "C0392B"),
        ("Retest Count", retest_cnt, "B9770E"),
    ]
    col = 2
    for label, value, color in kpis:
        ws.cell(row=5, column=col, value=label).font = Font(size=10, color="808080")
        vcell = ws.cell(row=6, column=col, value=value)
        vcell.font = Font(size=20, bold=True, color=color)
        col += 2

    r = 9
    ws.cell(row=r, column=2, value="Severity Breakdown — how wrong (% error)").font = Font(size=12, bold=True)
    r += 1
    for i, h in enumerate(["Severity", "Count", "% of Total", "Error % Range"]):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r += 1
    severity_ranges = {"PASS": "0-15%", "LOW": "15-30%", "MEDIUM": "30-60%", "HIGH": "60-100%"}
    for sev in ["HIGH", "MEDIUM", "LOW", "PASS"]:
        cnt = severity_counts.get(sev, 0)
        pct = (cnt / total * 100) if total else 0
        row_fill = SEVERITY_FILLS.get(sev)
        c0 = ws.cell(row=r, column=2, value=sev)
        c0.fill = row_fill
        c0.font = Font(bold=True, color=SEVERITY_FONT_COLORS.get(sev))
        ws.cell(row=r, column=3, value=cnt)
        ws.cell(row=r, column=4, value=f"{pct:.1f}%")
        ws.cell(row=r, column=5, value=severity_ranges.get(sev, ""))
        r += 1

    r += 2
    ws.cell(row=r, column=2, value="Domain Breakdown — where the errors are (sorted by error rate)").font = Font(size=12, bold=True)
    r += 1
    for i, h in enumerate(["Domain", "Total", "Pass", "Fail", "Retest", "Fail+Retest Rate %"]):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    r += 1

    rows = []
    for domain, stats in domain_stats.items():
        t = stats["total"]
        err = stats["fail"] + stats["retest"]
        rate = (err / t * 100) if t else 0
        rows.append((domain, stats["total"], stats["pass"], stats["fail"], stats["retest"], rate))
    rows.sort(key=lambda x: (-x[5], -(x[3] + x[4])))

    for domain, total_d, p, f, rt, rate in rows:
        ws.cell(row=r, column=2, value=domain)
        ws.cell(row=r, column=3, value=total_d)
        ws.cell(row=r, column=4, value=p).font = Font(color="1E8449")
        ws.cell(row=r, column=5, value=f).font = Font(color="C0392B")
        ws.cell(row=r, column=6, value=rt).font = Font(color="B9770E")
        rate_cell = ws.cell(row=r, column=7, value=round(rate, 1))
        if rate >= 40:
            rate_cell.font = Font(bold=True, color="C0392B")
        elif rate >= 15:
            rate_cell.font = Font(bold=True, color="B9770E")
        else:
            rate_cell.font = Font(bold=True, color="1E8449")
        r += 1

    for col_letter, width in zip("BCDEFG", [30, 10, 10, 10, 10, 18]):
        ws.column_dimensions[col_letter].width = width
