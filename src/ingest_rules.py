import os
import sys
from pathlib import Path
from typing import List

# Ensure project root and site-packages are in sys.path
project_root = Path(__file__).resolve().parent.parent
site_packages = project_root / "myenv" / "lib" / "python3.14" / "site-packages"
if site_packages.exists() and str(site_packages) not in sys.path:
    sys.path.insert(0, str(site_packages))
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import openpyxl
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import Chroma

import src.config as config


TEST_CASE_INDICATORS = [
    "evaluated", "kết quả", "ket qua", "testcase", "test_case", "result",
    "actual_resp", "vivi_listen", "auto_eval_result", "root_cause_analysis"
]


def is_test_result_file(file_path: Path) -> bool:
    """Checks if a file is a test execution log or evaluated output workbook rather than a rule spec."""
    lower_name = file_path.name.lower()
    if any(ind in lower_name for ind in ["evaluated", "kết quả", "ket qua", "testcase", "test_case", "result"]):
        return True
    return False


def parse_excel_rules(file_path: Path) -> List[Document]:
    """Parses Excel rule workbooks (e.g. command lists, error codes, preconditions)."""
    docs = []
    if not file_path.exists() or is_test_result_file(file_path):
        return docs

    print(f"📖 Parsing Excel rules file: {file_path.name}")
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"⚠️ Could not load Excel file {file_path}: {e}")
        return docs

    for sheet_name in wb.sheetnames:
        if sheet_name in ["Update_History", "(1) Readme"]:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        header = None
        start_row_idx = 0
        for r_idx, row in enumerate(rows[:5]):
            non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
            if len(non_empty) >= 2:
                header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row)]
                start_row_idx = r_idx + 1
                break

        if not header:
            continue

        # Skip sheet if it represents a test case result log
        header_lower_str = " ".join(header).lower()
        if any(k in header_lower_str for k in ["actual_resp", "vivi_listen", "auto_eval_result", "name_testcase", "latency"]):
            print(f"  ⏭️ Skipping test result sheet: '{sheet_name}' in {file_path.name}")
            continue

        for r_idx, row in enumerate(rows[start_row_idx:], start_row_idx + 1):
            if not any(c is not None and str(c).strip() for c in row):
                continue

            row_dict = {}
            for i, val in enumerate(row):
                if i < len(header):
                    col_name = header[i] or f"col_{i}"
                    row_dict[col_name] = str(val).strip() if val is not None else ""

            parts = [f"Rule Category/Domain: {sheet_name}"]
            rule_id = row_dict.get("ID", "") or row_dict.get("Error Code", "")
            if rule_id:
                parts.append(f"Rule ID: {rule_id}")

            for k, v in row_dict.items():
                if v and k not in ["col_0", "col_1", "col_2"]:
                    parts.append(f"{k}: {v}")

            text_content = "\n".join(parts)
            if len(text_content.strip()) > 20:
                docs.append(
                    Document(
                        page_content=text_content,
                        metadata={
                            "source": file_path.name,
                            "sheet": sheet_name,
                            "rule_id": rule_id,
                            "type": "excel_rule",
                        },
                    )
                )

    print(f"  └─ Extracted {len(docs)} rule items from {file_path.name}")
    return docs


def parse_markdown_rules(rules_dir: Path) -> List[Document]:
    """Parses markdown or text rule documents from rules directory."""
    docs = []
    if not rules_dir.exists():
        return docs

    for file_path in rules_dir.glob("**/*"):
        if file_path.suffix.lower() in [".md", ".txt"]:
            try:
                content = file_path.read_text(encoding="utf-8")
                if content.strip():
                    docs.append(
                        Document(
                            page_content=content,
                            metadata={"source": file_path.name, "type": "markdown_rule"},
                        )
                    )
            except Exception as e:
                print(f"⚠️ Error reading {file_path}: {e}")

    return docs


def ingest_rules():
    """Main function to ingest rules from Excel command lists and Markdown rules into ChromaDB."""
    all_rule_docs = []

    # 1. Ingest Excel rules from data/ directory (excluding test execution/evaluation files)
    data_path = Path(config.DATA_DIR)
    for xlsx_file in data_path.glob("*.xlsx"):
        if not xlsx_file.name.startswith("~$") and not is_test_result_file(xlsx_file):
            all_rule_docs.extend(parse_excel_rules(xlsx_file))

    # 2. Ingest Markdown/Text/Excel rules from data/rules/ directory
    rules_path = Path(config.RULES_DIR)
    rules_path.mkdir(parents=True, exist_ok=True)

    for xlsx_file in rules_path.glob("*.xlsx"):
        if not xlsx_file.name.startswith("~$") and not is_test_result_file(xlsx_file):
            all_rule_docs.extend(parse_excel_rules(xlsx_file))

    md_docs = parse_markdown_rules(rules_path)
    all_rule_docs.extend(md_docs)

    if not all_rule_docs:
        print("⚠️ No valid rule specification documents found in data/ or data/rules/.")
        return None

    print(f"Loaded {len(all_rule_docs)} raw rule document blocks.")

    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=config.CHUNK_SIZE,
        chunk_overlap=config.CHUNK_OVERLAP,
        length_function=len,
    )
    chunks = text_splitter.split_documents(all_rule_docs)
    print(f"Split rules into {len(chunks)} searchable chunks.")

    print(f"Generating embeddings and indexing into Rules ChromaDB at '{config.RULES_CHROMA_DB_DIR}'...")
    embeddings = config.get_embedding_model()

    vector_db = Chroma.from_documents(
        documents=chunks,
        embedding=embeddings,
        persist_directory=config.RULES_CHROMA_DB_DIR,
    )
    print(f"✅ Successfully indexed {len(chunks)} rule vectors in '{config.RULES_CHROMA_DB_DIR}'.")
    return vector_db


if __name__ == "__main__":
    ingest_rules()
