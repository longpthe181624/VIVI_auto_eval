import os
import re
import sys
import shutil
import openpyxl
from pathlib import Path
from typing import List

# Ensure project root and myenv site-packages are in sys.path
project_root = Path(__file__).resolve().parent.parent
site_packages = project_root / "myenv" / "lib" / "python3.14" / "site-packages"
if site_packages.exists() and str(site_packages) not in sys.path:
    sys.path.insert(0, str(site_packages))
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

try:
    from langchain_core.documents import Document
except ImportError:
    from langchain.docstore.document import Document

try:
    from langchain_community.document_loaders import DirectoryLoader, TextLoader, PyPDFLoader
except ImportError:
    from langchain.document_loaders import DirectoryLoader, TextLoader, PyPDFLoader

try:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
except ImportError:
    try:
        from langchain.text_splitter import RecursiveCharacterTextSplitter
    except ImportError:
        from langchain_classic.text_splitter import RecursiveCharacterTextSplitter

try:
    from langchain_community.vectorstores import Chroma
except ImportError:
    from langchain.vectorstores import Chroma

import src.config as config
from src.rag_chain import reset_vector_store


def is_test_result_file(file_name: str) -> bool:
    """Returns True if the file name represents a test execution output or log file."""
    fn = file_name.lower()
    return any(k in fn for k in ["evaluated_", "kết quả", "ket qua", "testcase", "result_"])


def load_text_documents(data_dir: str) -> List[Document]:
    """Loads all text, markdown, and PDF documents from data_dir and tags them with doc_type metadata."""
    data_path = Path(data_dir)
    if not data_path.exists():
        os.makedirs(data_path, exist_ok=True)
        return []

    documents = []
    
    # Load TXT and MD files
    txt_loader = DirectoryLoader(
        data_dir,
        glob="**/*.txt",
        loader_cls=TextLoader,
        loader_kwargs={"encoding": "utf-8"},
        silent_errors=True,
        show_progress=False,
    )
    for doc in txt_loader.load():
        src = str(doc.metadata.get("source", ""))
        fn = Path(src).name
        if "om_manuals" in src or "VinFast_OM_" in fn:
            doc.metadata["doc_type"] = "owner_manual"
            # Extract fine-grained metadata from filename e.g. VinFast_OM_VF8_2024_vi.txt
            match = re.search(r"VinFast_OM_([^_]+)_([^_]+)_([^\.]+)\.txt", fn)
            if match:
                doc.metadata["car_model"] = match.group(1)
                doc.metadata["version"] = match.group(2)
                doc.metadata["lang"] = match.group(3)
        else:
            doc.metadata["doc_type"] = "knowledge"
        documents.append(doc)

    md_loader = DirectoryLoader(
        data_dir,
        glob="**/*.md",
        loader_cls=TextLoader,
        loader_kwargs={"encoding": "utf-8"},
        silent_errors=True,
        show_progress=False,
    )
    for doc in md_loader.load():
        doc.metadata["doc_type"] = "knowledge"
        documents.append(doc)

    # Load PDF files
    pdf_loader = DirectoryLoader(
        data_dir,
        glob="**/*.pdf",
        loader_cls=PyPDFLoader,
        silent_errors=True,
        show_progress=False,
    )
    for doc in pdf_loader.load():
        doc.metadata["doc_type"] = "knowledge"
        documents.append(doc)

    return documents


def load_excel_rule_documents(data_dir: str) -> List[Document]:
    """Extracts command rules and domain specs from Excel workbooks in data_dir recursively."""
    data_path = Path(data_dir)
    docs = []

    for file_path in data_path.rglob("*.xlsx"):
        if is_test_result_file(file_path.name):
            print(f"  ⏭️ Skipping test result workbook: {file_path.name}")
            continue

        print(f"📖 Parsing Excel rules file: {file_path.name}")
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            print(f"  ⚠️ Could not load Excel file {file_path.name}: {e}")
            continue

        file_docs_count = 0
        for sheet_name in wb.sheetnames:
            if sheet_name in ["Update_History", "(1) Readme", "Summary"]:
                continue

            ws = wb[sheet_name]
            header = None
            consecutive_empty = 0
            sheet_rows = []

            for row in ws.iter_rows(values_only=True):
                if not any(c is not None and str(c).strip() for c in row):
                    consecutive_empty += 1
                    if consecutive_empty > 30:
                        break
                    continue

                consecutive_empty = 0
                sheet_rows.append(row)

            if len(sheet_rows) < 2:
                continue

            start_row_idx = 0
            for r_idx, row in enumerate(sheet_rows[:5]):
                non_empty = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if len(non_empty) >= 2:
                    header = [str(c).strip() if c is not None else f"col_{i}" for i, c in enumerate(row)]
                    start_row_idx = r_idx + 1
                    break

            if not header:
                continue

            header_lower_str = " ".join(header).lower()
            if any(k in header_lower_str for k in ["actual_resp", "vivi_listen", "auto_eval_result", "name_testcase", "latency"]):
                continue

            for row in sheet_rows[start_row_idx:]:
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(header):
                        col_name = header[i] or f"col_{i}"
                        row_dict[col_name] = str(val).strip() if val is not None else ""

                parts = [f"Rule Category/Domain: {sheet_name}"]
                rule_id = row_dict.get("ID", "") or row_dict.get("Error Code", "")
                if rule_id:
                    parts.append(f"Rule ID: {rule_id}")

                for k, v in row_dict.items():
                    if v and k not in ["col_0", "col_1", "col_2"]:
                        parts.append(f"{k}: {v}")

                text_content = "\n".join(parts)
                if len(text_content.strip()) > 20:
                    docs.append(
                        Document(
                            page_content=text_content,
                            metadata={
                                "source": file_path.name,
                                "sheet": sheet_name,
                                "rule_id": rule_id,
                                "doc_type": "command_rule",
                            },
                        )
                    )
                    file_docs_count += 1

        print(f"  └─ Extracted {file_docs_count} rule items from {file_path.name}")

    return docs


def ingest_documents():
    """Main execution function to load, split, embed, and store ALL documents & rules in 1 single ChromaDB collection."""
    print(f"\n============================================================")
    print(f"🚀 UNIFIED INGESTION ENGINE: Indexing Knowledge & Rules")
    print(f"============================================================")
    print(f"📂 Loading data from: {config.DATA_DIR}")

    # 1. Load Text & PDF documents (Owner Manuals, Knowledge)
    raw_text_docs = load_text_documents(config.DATA_DIR)
    
    # 2. Load Excel Command Rules
    rule_docs = load_excel_rule_documents(config.DATA_DIR)

    # 3. Split Text/Manual documents into semantic chunks
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=config.CHUNK_SIZE,
        chunk_overlap=config.CHUNK_OVERLAP,
        length_function=len,
    )

    all_chunks = []
    if raw_text_docs:
        manual_chunks = text_splitter.split_documents(raw_text_docs)
        print(f"✂️ Split {len(raw_text_docs)} manual/text files into {len(manual_chunks)} chunks.")
        all_chunks.extend(manual_chunks)

    if rule_docs:
        print(f"📋 Adding {len(rule_docs)} structured command rule documents.")
        all_chunks.extend(rule_docs)

    if not all_chunks:
        print("⚠️ No valid documents found in data directory.")
        return None

    print(f"📊 Total Combined Chunks to Index: {len(all_chunks)}")

    db_path = Path(config.CHROMA_DB_DIR)
    db_path.mkdir(parents=True, exist_ok=True)

    # 4. Embed & Persist into 1 single ChromaDB
    print("Generating embeddings and persisting vectors into 1 unified database...")
    embeddings = config.get_embedding_model()

    vector_db = Chroma.from_documents(
        documents=all_chunks,
        embedding=embeddings,
        persist_directory=config.CHROMA_DB_DIR,
    )

    reset_vector_store()

    print("\n" + "=" * 60)
    print(f"🎉 UNIFIED INGESTION COMPLETE!")
    print(f"💾 Stored {len(all_chunks)} vectors in single DB directory: '{config.CHROMA_DB_DIR}'")
    print("=" * 60)
    return vector_db


def ingest_single_file(file_path: Path, category: str = "kb"):
    """Incrementally embeds and adds a single file to ChromaDB without deleting existing DB."""
    file_path = Path(file_path)
    if not file_path.exists():
        print(f"⚠️ File '{file_path}' does not exist.")
        return None

    print(f"🚀 Incremental Ingestion: Processing '{file_path.name}' in category '{category}'...")
    docs = []
    if file_path.suffix.lower() == ".xlsx":
        if is_test_result_file(file_path.name):
            print(f"  ⏭️ Skipping test result workbook: {file_path.name}")
            return None
        rule_docs = load_excel_rule_documents(str(file_path.parent))
        docs = [d for d in rule_docs if d.metadata.get("source") == file_path.name]
    elif file_path.suffix.lower() in [".txt", ".md"]:
        text_loader = TextLoader(str(file_path), encoding="utf-8")
        raw_docs = text_loader.load()
        for doc in raw_docs:
            doc.metadata["source"] = file_path.name
            doc.metadata["doc_type"] = category
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=config.CHUNK_SIZE,
            chunk_overlap=config.CHUNK_OVERLAP,
            length_function=len,
        )
        docs = text_splitter.split_documents(raw_docs)
    elif file_path.suffix.lower() == ".pdf":
        pdf_loader = PyPDFLoader(str(file_path))
        raw_docs = pdf_loader.load()
        for doc in raw_docs:
            doc.metadata["source"] = file_path.name
            doc.metadata["doc_type"] = category
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=config.CHUNK_SIZE,
            chunk_overlap=config.CHUNK_OVERLAP,
            length_function=len,
        )
        docs = text_splitter.split_documents(raw_docs)

    if not docs:
        print(f"⚠️ No valid document chunks extracted from '{file_path.name}'.")
        return None

    db_path = Path(config.CHROMA_DB_DIR)
    db_path.mkdir(parents=True, exist_ok=True)
    embeddings = config.get_embedding_model()

    vector_db = Chroma(
        persist_directory=config.CHROMA_DB_DIR,
        embedding_function=embeddings,
    )

    try:
        if hasattr(vector_db, "_collection"):
            vector_db._collection.delete(where={"source": file_path.name})
    except Exception as e:
        print(f"  ℹ️ Note deleting old vectors for '{file_path.name}': {e}")

    vector_db.add_documents(docs)
    reset_vector_store()
    print(f"✅ Successfully incrementally embedded {len(docs)} chunks for '{file_path.name}' into ChromaDB.")
    return vector_db


def delete_single_file_vectors(source_name: str):
    """Purges vector chunks for source_name from ChromaDB without wiping the database directory."""
    db_path = Path(config.CHROMA_DB_DIR)
    if not db_path.exists():
        return
    try:
        embeddings = config.get_embedding_model()
        vector_db = Chroma(
            persist_directory=config.CHROMA_DB_DIR,
            embedding_function=embeddings,
        )
        if hasattr(vector_db, "_collection"):
            vector_db._collection.delete(where={"source": source_name})
        reset_vector_store()
        print(f"🗑️ Successfully purged vectors for '{source_name}' from ChromaDB.")
    except Exception as e:
        print(f"⚠️ Error deleting vectors for '{source_name}': {e}")


if __name__ == "__main__":
    ingest_documents()
