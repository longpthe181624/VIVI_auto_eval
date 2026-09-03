import os
import sys
import shutil
import uuid
import time
import asyncio
from pathlib import Path
from typing import Dict, Any, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

# Project root setup
project_root = Path(__file__).resolve().parent.parent
site_packages = project_root / "myenv" / "lib" / "python3.14" / "site-packages"
if site_packages.exists() and str(site_packages) not in sys.path:
    sys.path.insert(0, str(site_packages))
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

import src.config as config
from src.excel_evaluator import ExcelTestEvaluator, detect_testcase_category, trim_trailing_empty, _detect_columns
from src.eval_router import AdaptiveEvalRouter
from src.rag_chain import RAGChatbot, get_vector_store
from src.ingest import ingest_documents, load_text_documents, load_excel_rule_documents, is_test_result_file, ingest_single_file, delete_single_file_vectors
from src.report_builder import (
    EXTENDED_HEADERS, build_extended_row, style_header_row, style_data_row,
    autosize_columns, build_summary_sheet, extract_domain,
)

app = FastAPI(title="VIVI Auto-Eval Desktop Server", version="2.0.0")

# CORS middleware setup for desktop webview
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Static directory setup
STATIC_DIR = project_root / "static"
if not STATIC_DIR.exists():
    STATIC_DIR.mkdir(parents=True, exist_ok=True)

# In-memory execution state for batch evaluations
EVAL_TASKS: Dict[str, Dict[str, Any]] = {}
executor = ThreadPoolExecutor(max_workers=4)


# Pydantic Request Models
class ChatRequest(BaseModel):
    query: str
    car_model: Optional[str] = None
    lang: Optional[str] = None


class SingleEvalRequest(BaseModel):
    name: Optional[str] = "TC_Single"
    user_cmd: str
    actual_resp: str
    expected_resp: Optional[str] = ""
    error_log: Optional[str] = ""


class DeleteSourcePayload(BaseModel):
    source_name: str
    category: str


# Helper functions
def get_category_folders() -> List[str]:
    data_path = Path(config.DATA_DIR)
    default_cats = ["om_manuals", "command_rules", "error_codes", "troubleshooting", "warranty", "kb"]
    if not data_path.exists():
        data_path.mkdir(parents=True, exist_ok=True)
    
    found = [d.name for d in data_path.iterdir() if d.is_dir() and not d.name.startswith(".")]
    for c in default_cats:
        if c not in found:
            (data_path / c).mkdir(parents=True, exist_ok=True)
            found.append(c)
    return sorted(found)


# -------------------------------------------------------------------
# BATCH EXCEL EVALUATION TASK RUNNER
# -------------------------------------------------------------------
def run_batch_evaluation_task(task_id: str, file_path: str):
    task = EVAL_TASKS[task_id]
    task["status"] = "RUNNING"
    start_time = time.time()

    try:
        evaluator = ExcelTestEvaluator()
        input_path = Path(file_path)
        
        import openpyxl
        wb_in = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)

        total_rows_to_process = 0
        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]
            for row in ws_in.iter_rows(values_only=True):
                if trim_trailing_empty(row):
                    total_rows_to_process += 1
        
        task["total_rows"] = max(1, total_rows_to_process - len(wb_in.sheetnames))
        
        eval_count = 0
        pass_cnt = 0
        fail_cnt = 0
        retest_cnt = 0
        results_preview = []
        severity_counts: Dict[str, int] = {"PASS": 0, "LOW": 0, "MEDIUM": 0, "HIGH": 0}
        domain_stats: Dict[str, Dict[str, int]] = {}

        new_headers = EXTENDED_HEADERS

        for sheet_name in wb_in.sheetnames:
            ws_in = wb_in[sheet_name]
            ws_out = wb_out.create_sheet(title=sheet_name)

            header_vals = None
            col_map = {}
            raw_rows_buffer = []

            for r_idx, row in enumerate(ws_in.iter_rows(values_only=True), 1):
                trimmed_row = trim_trailing_empty(row)
                if not trimmed_row:
                    continue

                if header_vals is None:
                    test_map = _detect_columns(trimmed_row)
                    if "user_command" in test_map or "actual_resp" in test_map:
                        header_vals = [str(c).strip() if c is not None else "" for c in trimmed_row]
                        col_map = test_map
                        continue
                    else:
                        ws_out.append(trimmed_row)
                else:
                    raw_rows_buffer.append(trimmed_row)

            if not header_vals:
                continue

            max_data_col = max(col_map.values()) if col_map else len(header_vals) - 1
            max_col_idx = min(len(header_vals), max_data_col + 5)
            clean_header = header_vals[:max_col_idx]

            col_start_idx = col_map["auto_eval_start"] + 1 if "auto_eval_start" in col_map else len(clean_header) + 1
            full_header = clean_header + new_headers
            ws_out.append(full_header)
            style_header_row(ws_out, ws_out.max_row, col_start_idx, len(full_header))

            file_category = detect_testcase_category(input_path.name, sheet_name)

            def process_one_row(item):
                r_idx, row_vals = item
                name_val = str(row_vals[col_map["name"]]).strip() if "name" in col_map and col_map["name"] < len(row_vals) and row_vals[col_map["name"]] is not None else f"TC_{r_idx}"
                user_cmd = str(row_vals[col_map["user_command"]]).strip() if "user_command" in col_map and col_map["user_command"] < len(row_vals) and row_vals[col_map["user_command"]] is not None else ""
                vivi_listen = str(row_vals[col_map["vivi_listen"]]).strip() if "vivi_listen" in col_map and col_map["vivi_listen"] < len(row_vals) and row_vals[col_map["vivi_listen"]] is not None else ""
                actual_resp = str(row_vals[col_map["actual_resp"]]).strip() if "actual_resp" in col_map and col_map["actual_resp"] < len(row_vals) and row_vals[col_map["actual_resp"]] is not None else ""
                expected_resp = str(row_vals[col_map["expected_resp"]]).strip() if "expected_resp" in col_map and col_map["expected_resp"] < len(row_vals) and row_vals[col_map["expected_resp"]] is not None else ""

                res = evaluator.evaluate_row_sync(
                    name=name_val,
                    user_cmd=user_cmd,
                    actual_resp=actual_resp,
                    expected_resp=expected_resp,
                    category=file_category,
                    vivi_listen=vivi_listen
                )
                return r_idx, row_vals, name_val, user_cmd, actual_resp, expected_resp, res

            # Process rows in parallel (16 workers, matching the CLI path) instead of
            # one at a time - previously this loop ran sequentially, and with the LLM
            # Judge escalation (~4.6s/case on RETEST rows) that made large files take
            # tens of minutes even on an idle machine. Progress is still updated
            # incrementally via as_completed(), not just when the whole batch finishes.
            items = list(enumerate(raw_rows_buffer, 1))
            rows_by_idx: Dict[int, Any] = {}
            with ThreadPoolExecutor(max_workers=16) as pool:
                futures = {pool.submit(process_one_row, item): item[0] for item in items}
                for future in as_completed(futures):
                    r_idx, row_vals, name_val, user_cmd, actual_resp, expected_resp, res = future.result()

                    status = res["auto_result"]
                    severity = res.get("severity", "MEDIUM")
                    eval_count += 1
                    severity_counts[severity] = severity_counts.get(severity, 0) + 1
                    if status == "PASS":
                        pass_cnt += 1
                    elif status == "FAIL":
                        fail_cnt += 1
                    else:
                        retest_cnt += 1

                    domain = extract_domain(name_val)
                    d_stats = domain_stats.setdefault(domain, {"total": 0, "pass": 0, "fail": 0, "retest": 0})
                    d_stats["total"] += 1
                    d_stats["pass" if status == "PASS" else "fail" if status == "FAIL" else "retest"] += 1

                    elapsed = max(0.1, time.time() - start_time)
                    task["evaluated_rows"] = eval_count
                    task["pass_count"] = pass_cnt
                    task["fail_count"] = fail_cnt
                    task["retest_count"] = retest_cnt
                    task["progress_pct"] = min(100.0, round((eval_count / task["total_rows"]) * 100, 1))
                    task["rows_per_sec"] = round(eval_count / elapsed, 1)

                    eval_row_data = {
                        "id": name_val,
                        "user_command": user_cmd,
                        "actual_resp": actual_resp,
                        "expected_resp": expected_resp,
                        "status": status,
                        "score": res["score"],
                        "rule_info": res["rule_info"],
                        "rca": res["rca"],
                        "remediation": res["remediation"],
                        "trace_id": res.get("trace_id", ""),
                        "severity": res.get("severity", "PASS"),
                        "semantic_error_pct": res.get("semantic_error_pct", 0.0),
                        "error_category": res.get("error_category", "NONE"),
                        "trace_log": res.get("trace_log", {})
                    }
                    if status in ["FAIL", "RETEST"] or len(results_preview) < 5000:
                        results_preview.append(eval_row_data)

                    rows_by_idx[r_idx] = (row_vals, res, severity)

            # Write rows to the output sheet in the original input order, not
            # completion order (parallel execution finishes rows out of order).
            for r_idx in sorted(rows_by_idx.keys()):
                row_vals, res, severity = rows_by_idx[r_idx]
                clean_row_vals = list(row_vals[:max_col_idx])
                if len(clean_row_vals) < max_col_idx:
                    clean_row_vals.extend([""] * (max_col_idx - len(clean_row_vals)))

                append_vals = build_extended_row(res)

                if "auto_eval_start" in col_map:
                    out_row = clean_row_vals[:col_map["auto_eval_start"]] + append_vals
                else:
                    out_row = clean_row_vals + append_vals

                ws_out.append(out_row)
                style_data_row(ws_out, ws_out.max_row, col_start_idx, len(out_row), severity)

            autosize_columns(ws_out, full_header)

        build_summary_sheet(wb_out, eval_count, pass_cnt, fail_cnt, retest_cnt, severity_counts, domain_stats)

        output_file_name = f"evaluated_{input_path.stem}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_file_path = Path(config.DATA_DIR) / output_file_name
        wb_out.save(output_file_path)

        task["status"] = "COMPLETED"
        task["output_file"] = output_file_name
        task["results"] = results_preview
        task["progress_pct"] = 100.0

    except Exception as e:
        task["status"] = "FAILED"
        task["error"] = str(e)


# -------------------------------------------------------------------
# REST API ENDPOINTS
# -------------------------------------------------------------------

@app.get("/api/embedding/trace")
def get_embedding_trace(limit: int = 100):
    """Returns embedding-model call statistics and recent call log for monitoring."""
    from src.embedding_trace import get_embedding_trace_summary, get_recent_calls
    return {
        "summary": get_embedding_trace_summary(),
        "recent_calls": get_recent_calls(limit=limit),
    }


@app.get("/api/data/categories")
def list_categories():
    """Returns list of available category subdirectories."""
    return {"categories": get_category_folders()}


@app.get("/api/data/sources")
def list_embedded_sources():
    """Queries ChromaDB and data directory to return embedded file inventory with chunk counts."""
    data_path = Path(config.DATA_DIR)
    sources_dict: Dict[str, Dict[str, Any]] = {}

    # Query ChromaDB collection metadata
    try:
        vectorstore = get_vector_store(force_reload=True)
        if vectorstore and hasattr(vectorstore, "_collection"):
            raw_res = vectorstore._collection.get(include=["metadatas"])
            metadatas = raw_res.get("metadatas", []) or []
            for meta in metadatas:
                if not meta:
                    continue
                src_raw = meta.get("source", "Unknown")
                src_name = Path(src_raw).name
                doc_type = meta.get("doc_type", "general")

                if src_name not in sources_dict:
                    sources_dict[src_name] = {
                        "source_name": src_name,
                        "category": doc_type,
                        "chunks": 0,
                        "status": "Indexed"
                    }
                sources_dict[src_name]["chunks"] += 1
    except Exception as e:
        print(f"Error querying ChromaDB metadata: {e}")

    # Also scan disk files in data/ subdirectories to populate folder categories
    if data_path.exists():
        for category_dir in data_path.iterdir():
            if category_dir.is_dir() and not category_dir.name.startswith(".") and category_dir.name != "temp_uploads":
                cat_name = category_dir.name
                for file_p in category_dir.glob("*.*"):
                    if is_test_result_file(file_p.name):
                        continue
                    if file_p.name in sources_dict:
                        sources_dict[file_p.name]["category"] = cat_name
                    else:
                        sources_dict[file_p.name] = {
                            "source_name": file_p.name,
                            "category": cat_name,
                            "chunks": 0,
                            "status": "Pending Index"
                        }

    return {"sources": list(sources_dict.values()), "total_sources": len(sources_dict)}


@app.post("/api/data/upload")
async def upload_data_source(
    file: UploadFile = File(...),
    category: str = Form("kb")
):
    """Uploads data file to selected category subfolder and triggers incremental vector ingestion."""
    data_path = Path(config.DATA_DIR) / category
    data_path.mkdir(parents=True, exist_ok=True)

    dest_path = data_path / file.filename
    with open(dest_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    # Trigger incremental single-file vector indexing in background
    executor.submit(ingest_single_file, dest_path, category)
    return {
        "message": f"Successfully uploaded '{file.filename}' to category '{category}'. Vector indexing started.",
        "filename": file.filename,
        "category": category
    }


@app.post("/api/data/delete")
def delete_data_source(payload: DeleteSourcePayload):
    """Deletes specified file from data/<category>/ and purges its vector embeddings."""
    source_name = payload.source_name
    category = payload.category

    # Delete disk file if present
    deleted_files = 0
    data_path = Path(config.DATA_DIR)
    for target in data_path.rglob(source_name):
        if target.is_file():
            target.unlink()
            deleted_files += 1

    # Incremental vector purge
    executor.submit(delete_single_file_vectors, source_name)

    return {
        "message": f"Source '{source_name}' deleted. Purged vector embeddings for this file.",
        "deleted_files": deleted_files
    }


@app.post("/api/eval/upload")
async def upload_eval_excel(
    file: UploadFile = File(...),
    background_tasks: BackgroundTasks = BackgroundTasks()
):
    """Accepts `.xlsx` upload and starts async background evaluation task."""
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only Excel (.xlsx) files are supported.")

    task_id = str(uuid.uuid4())
    temp_dir = Path(config.DATA_DIR) / "temp_uploads"
    temp_dir.mkdir(parents=True, exist_ok=True)

    file_path = temp_dir / f"{task_id}_{file.filename}"
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    EVAL_TASKS[task_id] = {
        "task_id": task_id,
        "filename": file.filename,
        "status": "QUEUED",
        "progress_pct": 0.0,
        "evaluated_rows": 0,
        "total_rows": 0,
        "pass_count": 0,
        "fail_count": 0,
        "retest_count": 0,
        "rows_per_sec": 0.0,
        "output_file": "",
        "results": [],
        "error": None
    }

    background_tasks.add_task(run_batch_evaluation_task, task_id, str(file_path))
    return {"task_id": task_id, "message": "Evaluation task started."}


@app.get("/api/eval/progress/{task_id}")
def get_eval_progress(task_id: str):
    if task_id not in EVAL_TASKS:
        raise HTTPException(status_code=404, detail="Evaluation task not found.")
    return EVAL_TASKS[task_id]


@app.get("/api/eval/download/{filename}")
def download_eval_report(filename: str):
    file_path = Path(config.DATA_DIR) / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found.")
    return FileResponse(path=file_path, filename=filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/chat")
def chat_rag_assistant(payload: ChatRequest):
    """Queries RAG chatbot with optional car model/language filters."""
    chatbot = RAGChatbot()
    res = chatbot.answer_question(payload.query)
    return {
        "answer": res["answer"],
        "sources": res["sources"],
        "car_model": payload.car_model,
        "lang": payload.lang
    }


@app.post("/api/eval/single")
def evaluate_single_test_case(payload: SingleEvalRequest):
    """Evaluates a single test case row via AdaptiveEvalRouter."""
    router = AdaptiveEvalRouter()
    res = router.evaluate(
        name=payload.name,
        user_cmd=payload.user_cmd,
        actual_resp=payload.actual_resp,
        expected_resp=payload.expected_resp or ""
    )
    return res


# Serve static web interface
app.mount("/", StaticFiles(directory=str(STATIC_DIR), html=True), name="static")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)
